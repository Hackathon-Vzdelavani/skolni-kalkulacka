from dataclasses import dataclass
from pathlib import Path

import openpyxl
import scrapy
import selenium
from scrapy.selector import Selector
import time
from components.logger import Logger
from components.helper import error_message, driver_file
from selenium import webdriver
import re


@dataclass
class Predmet:
    nazev: str
    cile: str
    pozadavky: str
    obsah: str
    # predpoklady_znalosti: str
    # predpoklady_dovednosti: str
    # predpoklady_zpusobilosti: str
    # vysledky_znalosti: str
    # vysledky_dovednosti: str
    # vysledky_zpusobilosti: str

    @staticmethod
    def parse(driver: selenium.webdriver.Chrome):
        nazev = driver.find_element_by_xpath('//*[@id="prohlizeniDetail"]/div/table[1]/tbody/tr[3]/td[1]').text
        cile = driver.find_element_by_xpath('//*[@id="prohlizeniDetail"]/div/table[2]/tbody/tr[2]/td').text
        pozadavky = driver.find_element_by_xpath('//*[@id="prohlizeniDetail"]/div/table[2]/tbody/tr[4]/td').text
        obsah = driver.find_element_by_xpath('//*[@id="prohlizeniDetail"]/div/table[2]/tbody/tr[6]/td').text

        return Predmet(
            nazev, cile, pozadavky, obsah)


class SkillSpider(scrapy.Spider):
    allowed_domains = ['www.zcu.cz']
    name = "predmety_spider"
    start_urls = ["https://portal.zcu.cz/portal/studium/prohlizeni.html?pc_phs=-2121444242&pc_windowid=615122&pc_publicnavigationalstatechanges=AAAAAA**&pc_phase=render&pc_type=portlet&pc_navigationalstate=JBPNS_rO0ABXdRAApzdGF0ZUNsYXNzAAAAAQA2Y3ouemN1LnN0YWcucG9ydGxldHMxNjgucHJvaGxpemVuaS5zdGF0ZXMuUHJlZG1ldFN0YXRlAAdfX0VPRl9f#prohlizeniContent"]

    def __init__(self):
        super().__init__()
        self.logging = Logger(spider=self.name).logger
        self.driver = webdriver.Chrome(driver_file())
        self.url_index = 0
        self.predmet_params = self.load_export_predmetu()
        self.param_index = 0

    def start_requests(self):
        for url in self.start_urls:
            print("Opening", len(self.predmet_params), "params")
            for _ in range(len(self.predmet_params))[:50]:
                yield scrapy.Request(url=url, callback=self.parse, dont_filter=True)

    def load_export_predmetu(self):
        sheet = openpyxl.load_workbook(Path("data/exportPredmetu.xlsx")).active
        max_row = sheet.max_row
        params = []
        for row_i in range(3, max_row+1):
            params.append([
                sheet.cell(row=row_i, column=1).value,
                sheet.cell(row=row_i, column=2).value,
            ])

        return params

    def parse(self, response):
        self.logging.info('Visiting ' + response.url)
        self.driver.get(response.url)
        pracoviste_field = self.driver.find_element_by_id("katedraInput")
        zkratka_field = self.driver.find_element_by_id("predmetInput")
        submit_button = self.driver.find_element_by_id("searchPredmetySubmit")

        pracoviste, zkratka = self.predmet_params[self.param_index]
        self.param_index += 1

        pracoviste_field.send_keys(pracoviste)
        zkratka_field.send_keys(zkratka)
        submit_button.click()

        # predemet loaded
        print(Predmet.parse(self.driver))


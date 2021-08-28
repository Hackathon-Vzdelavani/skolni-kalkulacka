from components.logger import Logger
from components.database import Database
import openpyxl
import os
FACULTIES = ["fav", "fdu", "fek", "ff", "fpe", "fpr", "fst", "fzs"]
parent_directory = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))

DICT_KEYS = ["stprIdno", "nazev", "kod", "titul", "titulZkr", "titulRigo", "titulRigoZkr", "typ", "forma", "fakulta",
             "stdDelka", "maxDelka", "kredity", "vykazovan", "platnyOd", "neplatnyOd", "cile", "garant", "garantUcitIdno",
             "garantAdministrace", "garantAdministraceUcitIdno", "jazyk", "profil", "akreditacePoNovele",
             "akreditaceInstitucionalni", "akreditaceOdDate", "akreditaceDoDate", "akreditaceZtracenaOdDate",
             "akreditovanSeSpecializaci", "akreditaceCislo", "kodIsced", "navaznostNaDalsiSpCz", "navaznostNaDalsiSpAn",
             "navaznostNaDalsiSpJ3", "navaznostNaDalsiSpJ4", "obsahoveZamereniCz", "obsahoveZamereniAn",
             "obsahoveZamereniJ3", "obsahoveZamereniJ4", "pozadovanaZdrZpusobilostCz", "pozadovanaZdrZpusobilostAn",
             "pozadovanaZdrZpusobilostJ3", "pozadovanaZdrZpusobilostJ4", "profilAbsolventaCz", "profilAbsolventaAn",
             "profilAbsolventaJ3", "profilAbsolventaJ4", "pocetPrijimanych", "pocetPrijimanychPoznamka",
             "pozadavkyNaPrijetiCz", "pozadavkyNaPrijetiAn", "pozadavkyNaPrijetiUrl", "pozadavkyRovnyPristupZdrCz",
             "pozadavkyRovnyPristupZdrAn", "pozadavkyRovnyPristupZdrJ3", "pozadavkyRovnyPristupZdrJ4",
             "predpokladUplatitelnostiCz", "predpokladUplatitelnostiAn", "predpokladUplatitelnostiJ3",
             "predpokladUplatitelnostiJ4", "regulovanePovolaniCz", "regulovanePovolaniAn", "regulovanePovolaniJ3",
             "regulovanePovolaniJ4", "regulovanePovolaniUznavaciO", "regulovanePovolaniDatum", "moznePracovniPoziceCz",
             "moznePracovniPoziceAn", "moznePracovniPoziceJ3", "moznePracovniPoziceJ4", "vyukaZahranicniPredpis",
             "vyukaAVCRPracoviste", "akreditaceOmezenaDuvod", "akreditaceOmezenaCislo", "akreditaceOmezenaOd",
             "koefEkonomickeNarocnosti", "typickaTemataVskp", "praktickaVyuka", "spolupraceSPraxi", "odbornaPraxe",
             "jointDegrees", "poznamkaVerejna", "podminkyPokracovaniStudia", "nazevProDiplomCz", "nazevProDiplomAn"]

class ProgramParser:
    allowed_domains = ['www.zcu.cz']
    def __init__(self):
        self.database = Database()

    def run(self):
        for faculty_name in FACULTIES:
            programs = self.open_excel_file(faculty_name)
            for program in programs:
                item = self.prepare_program_item(program)
                self.database.insert_program_detail(item)

    def prepare_program_item(self, data):
        print(data)
        return {
            "id": data["stprIdno"],
            "name": data["nazev"],
            "code": data["kod"],
            "title": data["titulZkr"],
            "type": data["typ"],
            "form": data["forma"],
            "faculty": data["fakulta"],
            "length": data["stdDelka"],
            "goal": data["cile"],
            "garant": data["garant"],
            "language": data["jazyk"],
        }


    def open_excel_file(self, faculty_name):
        file_name = f"programy_{faculty_name}.xlsx"
        excel_file_path = os.path.join(parent_directory, "data", file_name)
        sheet = openpyxl.load_workbook(excel_file_path).active
        programs = []
        for row_index in range(2, sheet.max_row+1):
            programs.append({column: sheet.cell(row=row_index, column=col_index+1).value for col_index, column in enumerate(DICT_KEYS)})
        return programs
